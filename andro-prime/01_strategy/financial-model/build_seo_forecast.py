"""
Build andro-prime-seo-channel-forecast.xlsx — 18-month SEO/content revenue model.

Standalone channel model (NOT a rewrite of the phase-0 P&L). Every output is a
live Excel formula referencing the Assumptions + Traffic_Ramp sheets, so Keith can
flex any lever (conversion %, AOV, the monthly session ramp, TRT-cohort conversion)
and the whole model recalculates.

Grounding (per session 2026-05-31):
- Search volumes / underserved clusters: 06_marketing/seo-ai-search/portfolio-demand-gap-map.md
- LTV / tenure: longitudinal-tracker-decision (3-mo £37.50 -> 12-mo £77.32), AOV from kit prices
- Honest caveats: flawless CONTENT execution != guaranteed rankings on a new domain;
  SEO is back-loaded; conversion + whether-rankings-come are the two swing variables.

Default scenarios (all editable):
- Traffic ramp hits these milestones (sessions/mo): see Traffic_Ramp sheet.
    Conservative  M6 1.5k / M12 8k  / M18 18k
    Base          M6 4k   / M12 20k / M18 40k
    Strong        M6 7k   / M12 32k / M18 70k
- Blog->kit conversion: Cons 0.3% / Base 0.5% / Strong 1.0%
- Kit AOV blended £135 ; supplement+retest uplift £25 -> eff. rev/customer £160
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"d:\Androprime_main\andro-prime\01_strategy\financial-model\andro-prime-seo-channel-forecast.xlsx"

# ---------------- Style helpers ----------------
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
CONS_FILL = PatternFill("solid", fgColor="FCE4D6")
BASE_FILL = PatternFill("solid", fgColor="E2EFDA")
STRONG_FILL = PatternFill("solid", fgColor="DDEBF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D6DCE4")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")
BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

GBP = u'£#,##0'
PCT = '0.0%'
INT = '#,##0'


def hdr(ws, row, values, start_col=1, fill=HDR_FILL, font=WHITE):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def cell(ws, r, col, v, fill=None, font=None, fmt=None, border=True):
    c = ws.cell(row=r, column=col, value=v)
    if fill: c.fill = fill
    if font: c.font = font
    if fmt: c.number_format = fmt
    if border: c.border = BORDER
    return c


def autosize(ws, min_w=11, max_w=46):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None: continue
            L = get_column_letter(c.column)
            widths[L] = min(max(widths.get(L, min_w), len(str(c.value)) + 2), max_w)
    for L, w in widths.items():
        ws.column_dimensions[L].width = w


# ---------------- Default inputs ----------------
MONTHS = list(range(1, 19))
# Back-loaded monthly session ramps that hit the milestone table above.
RAMP_CONS   = [120, 280, 480, 720, 1050, 1500, 2100, 2900, 3900, 5200, 6500, 8000, 9800, 11500, 13000, 15000, 16500, 18000]
RAMP_BASE   = [200, 500, 900, 1500, 2500, 4000, 6000, 8500, 12000, 15000, 18000, 20000, 24000, 28000, 32000, 35000, 38000, 40000]
RAMP_STRONG = [400, 1000, 1800, 3000, 5000, 7000, 10000, 14000, 18000, 23000, 28000, 32000, 40000, 48000, 55000, 60000, 65000, 70000]

wb = Workbook()

# =========================================================
# Sheet 1 — Assumptions (editable levers)
# =========================================================
a = wb.active
a.title = "Assumptions"
cell(a, 1, 1, "SEO CHANNEL FORECAST — ASSUMPTIONS (edit yellow cells; everything else recalculates)",
     font=Font(bold=True, size=12), border=False)
a.merge_cells("A1:D1")

hdr(a, 3, ["Lever", "Value", "Unit", "Note"])
rows = [
    ("Kit AOV (blended)", 135, GBP, "Mix of Kit 1 £99 / Kit 2 £119 / Kit 3 £179 / Kit 3 Plus £239"),
    ("Supplement + retest uplift / customer", 25, GBP, "~20-30% attach; LTV £37.50 (3-mo) -> £77.32 (12-mo tenure)"),
]
r = 4
for name, val, fmt, note in rows:
    cell(a, r, 1, name, font=BOLD)
    cell(a, r, 2, val, fill=INPUT_FILL, fmt=fmt)
    cell(a, r, 3, "GBP")
    cell(a, r, 4, note, font=ITALIC)
    r += 1
cell(a, r, 1, "Effective revenue / kit customer", font=BOLD)
cell(a, r, 2, "=B4+B5", fmt=GBP, fill=TOTAL_FILL, font=BOLD)
cell(a, r, 4, "AOV + supplement/retest uplift (18-mo window)", font=ITALIC)
EFFREV = f"Assumptions!$B${r}"
r += 2

hdr(a, r, ["Blog -> kit conversion", "Rate", "", "Note"])
conv_row0 = r + 1
for i, (name, val, note) in enumerate([
    ("Conservative", 0.003, "Mostly informational traffic; weak funnel"),
    ("Base", 0.005, "Semi-commercial 'X blood test' terms + CIO nurture"),
    ("Strong", 0.010, "High commercial-intent mix + tight CRO"),
]):
    rr = conv_row0 + i
    cell(a, rr, 1, name, font=BOLD)
    cell(a, rr, 2, val, fill=INPUT_FILL, fmt=PCT)
    cell(a, rr, 4, note, font=ITALIC)
CONV = {"cons": f"Assumptions!$B${conv_row0}",
        "base": f"Assumptions!$B${conv_row0+1}",
        "strong": f"Assumptions!$B${conv_row0+2}"}
r = conv_row0 + 4

hdr(a, r, ["Downstream CQC/TRT cohort (post-18mo prize)", "Value", "", "Note"])
d0 = r + 1
dn = [
    ("TRT conversion of wellness cohort", 0.12, PCT, "Share of kit customers -> TRT after CQC"),
    ("TRT ARPU / month", 200, GBP, "Model: men pay £185-220/mo new-brand TRT"),
    ("Billed months (Year-1 of TRT)", 12, INT, "First 12 months of the TRT subscription"),
]
for i, (name, val, fmt, note) in enumerate(dn):
    rr = d0 + i
    cell(a, rr, 1, name, font=BOLD)
    cell(a, rr, 2, val, fill=INPUT_FILL, fmt=fmt)
    cell(a, rr, 4, note, font=ITALIC)
TRT_CONV, TRT_ARPU, TRT_MO = f"Assumptions!$B${d0}", f"Assumptions!$B${d0+1}", f"Assumptions!$B${d0+2}"
autosize(a)

# =========================================================
# Sheet 2 — Traffic_Ramp (editable monthly sessions)
# =========================================================
t = wb.create_sheet("Traffic_Ramp")
cell(t, 1, 1, "MONTHLY ORGANIC SESSIONS (edit any cell; back-loaded by design — SEO matures M12-18)",
     font=Font(bold=True, size=12), border=False)
t.merge_cells("A1:D1")
hdr(t, 3, ["Month", "Conservative", "Base", "Strong"])
ramp_row0 = 4
for i, m in enumerate(MONTHS):
    rr = ramp_row0 + i
    cell(t, rr, 1, f"M{m}", font=BOLD)
    cell(t, rr, 2, RAMP_CONS[i], fill=INPUT_FILL, fmt=INT)
    cell(t, rr, 3, RAMP_BASE[i], fill=INPUT_FILL, fmt=INT)
    cell(t, rr, 4, RAMP_STRONG[i], fill=INPUT_FILL, fmt=INT)
tot = ramp_row0 + len(MONTHS)
cell(t, tot, 1, "18-mo total", font=BOLD, fill=TOTAL_FILL)
for col in (2, 3, 4):
    L = get_column_letter(col)
    cell(t, tot, col, f"=SUM({L}{ramp_row0}:{L}{tot-1})", font=BOLD, fill=TOTAL_FILL, fmt=INT)
autosize(t)

# =========================================================
# Sheet 3 — Forecast (formula-driven)
# =========================================================
f = wb.create_sheet("Forecast")
cell(f, 1, 1, "18-MONTH SEO REVENUE FORECAST (all formulas — driven by Assumptions + Traffic_Ramp)",
     font=Font(bold=True, size=12), border=False)
f.merge_cells("A1:J1")
hdr(f, 2, ["", "CONSERVATIVE", "", "", "BASE", "", "", "STRONG", "", ""])
for c0, fill in ((2, CONS_FILL), (5, BASE_FILL), (8, STRONG_FILL)):
    for off in range(3):
        f.cell(row=2, column=c0+off).fill = fill
hdr(f, 3, ["Month", "Sessions", "Customers", "Revenue",
           "Sessions", "Customers", "Revenue",
           "Sessions", "Customers", "Revenue"])
fr0 = 4
for i, m in enumerate(MONTHS):
    rr = fr0 + i
    tr = ramp_row0 + i  # matching Traffic_Ramp row
    cell(f, rr, 1, f"M{m}", font=BOLD)
    for (scol, tcol, conv, fill) in [(2, "B", CONV["cons"], CONS_FILL),
                                     (5, "C", CONV["base"], BASE_FILL),
                                     (8, "D", CONV["strong"], STRONG_FILL)]:
        sess = f"Traffic_Ramp!{tcol}{tr}"
        cell(f, rr, scol, f"={sess}", fmt=INT)
        cust_cell = f"{get_column_letter(scol)}{rr}"
        cell(f, rr, scol+1, f"={sess}*{conv}", fmt='#,##0.0')
        cell(f, rr, scol+2, f"={get_column_letter(scol+1)}{rr}*{EFFREV}", fmt=GBP, fill=fill)
ft = fr0 + len(MONTHS)
cell(f, ft, 1, "18-MO TOTAL", font=BOLD, fill=TOTAL_FILL)
for col in range(2, 11):
    L = get_column_letter(col)
    fmt = GBP if col in (4, 7, 10) else ('#,##0.0' if col in (3, 6, 9) else INT)
    cell(f, ft, col, f"=SUM({L}{fr0}:{L}{ft-1})", font=BOLD, fill=TOTAL_FILL, fmt=fmt)
# Run-rate row (M18)
rrow = ft + 1
m18 = fr0 + 17
cell(f, rrow, 1, "M18 run-rate (orders/wk)", font=BOLD)
for col in (3, 6, 9):
    L = get_column_letter(col)
    cell(f, rrow, col+1, f"={L}{m18}/4.33", fmt='#,##0.0', font=BOLD)
autosize(f)

# =========================================================
# Sheet 4 — Downstream (the real prize)
# =========================================================
d = wb.create_sheet("Downstream")
cell(d, 1, 1, "WHY THE 18-MONTH NUMBER UNDERSTATES IT (the content is feeding these)",
     font=Font(bold=True, size=12), border=False)
d.merge_cells("A1:E1")
notes = [
    "1. COMPOUNDING TRAFFIC ASSET",
    "   M18 sessions are the launch pad for Year 2-3; same content typically 2-3x's with no extra spend.",
    "   The 18-mo revenue is the bottom of the curve, not the plateau.",
    "",
    "2. PRE-QUALIFIED CQC / TRT COHORT  (quantified below)",
    "   Every kit customer is a warm lead for TRT at ~80% margin post-CQC. Most of this lands JUST past 18 months.",
    "",
    "3. AI-CITATION UPSIDE (Pillar F)  — unmodellable; if the original-data + Ewa E-E-A-T play lands it's a step-change.",
]
rr = 3
for n in notes:
    cell(d, rr, 1, n, font=(BOLD if n and n[0].isdigit() else ITALIC), border=False)
    rr += 1

rr += 1
hdr(d, rr, ["Downstream TRT value from the 18-mo cohort", "Conservative", "Base", "Strong"])
cust_total_row = ft
trt_row = rr + 1
cell(d, trt_row, 1, "18-mo kit customers (from Forecast)", font=BOLD)
for i, col in enumerate(("C", "F", "I")):  # customers totals on Forecast
    cell(d, trt_row, 2+i, f"=Forecast!{col}{cust_total_row}", fmt='#,##0')
cell(d, trt_row+1, 1, "-> annual TRT revenue if cohort converts", font=BOLD)
for i in range(3):
    L = get_column_letter(2+i)
    cell(d, trt_row+1, 2+i, f"={L}{trt_row}*{TRT_CONV}*{TRT_ARPU}*{TRT_MO}", fmt=GBP, fill=TOTAL_FILL, font=BOLD)
cell(d, trt_row+3, 1,
     "Illustrative annual recurring TRT revenue the wellness cohort seeds (lands post-CQC, mostly beyond month 18).",
     font=ITALIC, border=False)
cell(d, trt_row+4, 1,
     "Compare to the 18-mo direct kit line on Forecast — the recurring clinical line is the larger prize.",
     font=ITALIC, border=False)
autosize(d)

# Order sheets
wb.move_sheet("Assumptions", -wb.sheetnames.index("Assumptions"))
wb.save(OUTPUT)
print("WROTE", OUTPUT)

# ---- print computed defaults for the chat report ----
eff = 135 + 25
for label, ramp, conv in [("Conservative", RAMP_CONS, 0.003), ("Base", RAMP_BASE, 0.005), ("Strong", RAMP_STRONG, 0.010)]:
    sess = sum(ramp)
    cust = sess * conv
    rev = cust * eff
    wk = ramp[-1] * conv / 4.33
    print(f"{label:13s} sessions={sess:>8,}  customers={cust:>7.0f}  revenue=£{rev:>9,.0f}  M18 run-rate={wk:>4.1f}/wk")
trt_arpu, trt_conv, trt_mo = 200, 0.12, 12
for label, ramp, conv in [("Conservative", RAMP_CONS, 0.003), ("Base", RAMP_BASE, 0.005), ("Strong", RAMP_STRONG, 0.010)]:
    cust = sum(ramp) * conv
    print(f"{label:13s} downstream annual TRT = £{cust*trt_conv*trt_arpu*trt_mo:>10,.0f}")
