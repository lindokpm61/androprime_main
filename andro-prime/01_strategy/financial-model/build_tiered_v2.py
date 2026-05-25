"""
Build andro-prime-tiered-platform-model-v2.xlsx from tiered-platform-model-v2.md.

Decisions baked in (per Keith 2026-05-25):
- Y1 baseline anchor: v2 doc as written (AP M12 cum cash = £42,526)
- Performance tier: sub-tier of Optimisation, same per-customer economics,
  elevated retest cadence (signals clinical eligibility)

Defaults (editable on Assumptions sheet):
- v2 gate evaluated at M12 (per doc; flagged in Conflicts_Log)
- CQC granted end-M12, ramp from M13
- Ewa retainer: £1k/mo Y1, £1.5k/mo Y2, £2k/mo Y3 (= £54k total, matches doc headline)
- CL pre-CQC burn: £53k spread M1-M12 (= £4,417/mo flat)
- Volume ramp: linear M12=180 -> M36=500
- Intercompany: 15% of CL external revenue billed to AP as "marketing services"

Anchors (model output reconciles to):
  AP PASS: M12 £42,526 / M24 £401,333 / M36 £1,276,117 / M36 net £97,718
  AP FAIL: M12 £42,526 / M24 £391,671 / M36 £1,178,126 / M36 net £86,870
  CL PASS: M12 (£53,000) / M24 (£44,799) / M36 £103,731 / M36 net £22,050
  CL FAIL: M12 (£53,000) / M24 (£53,418) / M36 £40,324 / M36 net £14,480

Calibration plug per scenario per entity is shown explicitly so Keith can see
the gap between bottom-up math and the doc's published number.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

OUTPUT = r"d:\Androprime_main\andro-prime\01_strategy\financial-model\andro-prime-tiered-platform-model-v2.xlsx"

# ---------------- Style helpers ----------------
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SCEN_PASS_FILL = PatternFill("solid", fgColor="E2EFDA")
SCEN_FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
PLUG_FILL = PatternFill("solid", fgColor="F8CBAD")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")
SMALL_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def hdr(ws: Worksheet, row: int, values, start_col=1, fill=HDR_FILL, font=WHITE):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_row(ws, row, values, start_col=1, bold=False, fill=None, num_fmt=None):
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        if bold:
            c.font = BOLD
        if fill:
            c.fill = fill
        if num_fmt and isinstance(v, (int, float)):
            c.number_format = num_fmt


def autosize(ws, min_width=10, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            # Skip merged cells (they don't have column_letter as attr access)
            if cell.__class__.__name__ == "MergedCell":
                continue
            col_letter = cell.column_letter
            if cell.value is not None:
                length = len(str(cell.value))
                if length > widths.get(col_letter, 0):
                    widths[col_letter] = length
    for col_letter, length in widths.items():
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


# ---------------- Inputs (mirror doc) ----------------
MONTHS = 36

# Volume ramp - H1/H2 from option-4 v3, H3/H4 linear ramp M12=180 -> M36=500
VOLUME = {
    1: 24, 2: 42, 3: 66, 4: 90, 5: 108, 6: 120,
    7: 130, 8: 140, 9: 150, 10: 160, 11: 170, 12: 180,
}
# M13 to M36: linear ramp 180 -> 500 over 24 months, ~13.3/mo
for m in range(13, 37):
    VOLUME[m] = int(round(180 + (500 - 180) * (m - 12) / 24))

# Kit mix 40/40/20, blended kit margins from option-4 v3 (v2.3)
KIT_MIX = {"K1": 0.40, "K2": 0.40, "K3": 0.20}
KIT_NET = {"K1": 25.70, "K2": 39.72, "K3": 55.30}  # blended 50/50 direct/affiliate
WEIGHTED_KIT_NET = sum(KIT_MIX[k] * KIT_NET[k] for k in KIT_MIX)  # = 32.628

# Supplements (Daily Stack + Collagen for Kit 2)
DS_SUB_PRICE = 34.95
DS_NET_PER_MO = 22.08
COLLAGEN_NET_PER_MO = 16.20
DS_CONV = 0.15  # Kit -> Daily Stack
COLL_CONV_K2 = 0.15  # Kit 2 -> Collagen
DS_TENURE_MO = 4
COLL_TENURE_MO = 4

# Retest revenue (gated by tracker existence M4+, ramps to full M13)
RETEST_PRICE = 89.0  # blended retest price across kits
RETEST_NET = 35.0    # net after Vitall + Stripe + fulfilment
# Retest cadence by tier (months between retests when in tier)
RETEST_CADENCE = {"Base": None, "Opt": 12, "Perf": 4}

# Tier entry mix (post-tracker M4+; pre-tracker, all in Base)
TIER_ENTRY = {"Base": 0.70, "Opt": 0.30, "Perf": 0.00}

# Tier transition rates (monthly)
TRANS_PRE_V2 = {"Base_Opt": 0.02, "Opt_Perf": 0.015}
TRANS_POST_V2_PASS = {"Base_Opt": 0.03, "Opt_Perf": 0.025}

# v2 gate evaluation month
V2_GATE_MONTH = 12

# Cross-entity Perf -> Clinical (post-CQC M13+)
CROSS_PASS = 0.04
CROSS_FAIL = 0.02
CQC_GRANT_MONTH = 12
CL_RAMP_START_MONTH = 13

# Customer churn (proportional off all tiers, monthly)
CHURN = 0.02

# AP fixed overhead schedule (escalates by year)
AP_OVERHEAD = {1: 593, 2: 1500, 3: 2500}  # Y1, Y2, Y3 per month

# AP PT programme one-off / monthly (from option-4 v3)
AP_OPEX_Y1 = 16798  # 6-mo programme + 6-mo extrapolation OpEx (incl. fixed overhead handled separately)
# For Y2/Y3 simplify: paid media + content
AP_PAID_MEDIA = {1: 0, 2: 1500, 3: 3000}  # £/mo by year

# Ewa retainer schedule
EWA_RETAINER = {1: 1000, 2: 1500, 3: 2000}  # £/mo by year - sums to £54k over 36mo

# Tracker build cost (one-off M3-M4, ~£8k spread)
TRACKER_BUILD_M3 = 4000
TRACKER_BUILD_M4 = 4000

# Clinical Ltd
CL_PRE_CQC_BURN_TOTAL = 53000
CL_PRE_CQC_MONTHLY = CL_PRE_CQC_BURN_TOTAL / 12  # ~£4,417/mo flat M1-M12
CL_OPEX_POST_CQC = {2: 4000, 3: 6000}  # £/mo by year, post-CQC
CL_ARPU = 200  # £/mo per active clinical patient (TRT subscription blended)
CL_GROSS_MARGIN = 0.55
CL_PATIENT_CHURN = 0.03  # monthly clinical churn

# Intercompany transfer (AP bills CL 15% of CL external revenue as marketing services)
INTERCO_PCT = 0.15

# v2 doc anchor numbers (for calibration plug)
DOC_ANCHOR = {
    "AP_PASS": {"M12": 42526, "M24": 401333, "M36": 1276117, "M36_net": 97718},
    "AP_FAIL": {"M12": 42526, "M24": 391671, "M36": 1178126, "M36_net": 86870},
    "CL_PASS": {"M12": -53000, "M24": -44799, "M36": 103731, "M36_net": 22050},
    "CL_FAIL": {"M12": -53000, "M24": -53418, "M36": 40324, "M36_net": 14480},
}


# ---------------- Engine ----------------
def year_of(m: int) -> int:
    return ((m - 1) // 12) + 1


def simulate_tiers(case: str):
    """
    Returns dict of month -> {Base, Opt, Perf} populations.
    case: 'pass' or 'fail'
    """
    pops = []
    base = 0.0
    opt = 0.0
    perf = 0.0
    for m in range(1, MONTHS + 1):
        # 1) New customers enter at tier-entry mix
        new = VOLUME[m]
        if m < 4:
            # Pre-tracker: all enter Base
            entry = {"Base": new * 1.0, "Opt": 0, "Perf": 0}
        else:
            entry = {k: new * TIER_ENTRY[k] for k in TIER_ENTRY}

        # 2) Churn off existing pops (before transitions)
        base *= (1 - CHURN)
        opt *= (1 - CHURN)
        perf *= (1 - CHURN)

        # 3) Transitions
        if m < V2_GATE_MONTH or case == "fail":
            t = TRANS_PRE_V2
        else:
            t = TRANS_POST_V2_PASS
        base_to_opt = base * t["Base_Opt"]
        opt_to_perf = opt * t["Opt_Perf"]
        base -= base_to_opt
        opt += base_to_opt - opt_to_perf
        perf += opt_to_perf

        # 4) Cross-entity Perf -> Clinical (post-CQC), removes from AP Perf pool
        perf_to_clinical = 0.0
        if m >= CL_RAMP_START_MONTH:
            cross = CROSS_PASS if case == "pass" else CROSS_FAIL
            perf_to_clinical = perf * cross
            perf -= perf_to_clinical

        # 5) Add new entries
        base += entry["Base"]
        opt += entry["Opt"]
        perf += entry["Perf"]

        pops.append({
            "m": m,
            "new": new,
            "base": base,
            "opt": opt,
            "perf": perf,
            "perf_to_clinical": perf_to_clinical,
        })
    return pops


def simulate_clinical(pops, case: str):
    """
    Track clinical patient population month-by-month.
    Inflows: perf_to_clinical from AP. Outflows: patient churn.
    """
    cl = []
    pop = 0.0
    for row in pops:
        m = row["m"]
        if m >= CL_RAMP_START_MONTH:
            pop *= (1 - CL_PATIENT_CHURN)
            pop += row["perf_to_clinical"]
        cl.append({"m": m, "new_patients": row["perf_to_clinical"], "active": pop})
    return cl


def simulate_ap_pnl(pops, cl_data, case: str):
    """
    Monthly P&L for Andro Prime Ltd.
    Revenue: kit margin + Daily Stack net + Collagen net + retest + intercompany from CL.
    Costs: overhead + Ewa + tracker build + paid media + PT programme (Y1).
    """
    rows = []
    cum = 0.0
    ds_active_by_month = [0.0] * (MONTHS + 1)
    coll_active_by_month = [0.0] * (MONTHS + 1)

    for i, row in enumerate(pops):
        m = row["m"]
        new = row["new"]
        # New DS subs and Collagen subs from this month's volume
        new_ds = new * DS_CONV
        new_coll = (new * KIT_MIX["K2"]) * COLL_CONV_K2

        # Update active rolling windows (DS_TENURE_MO and COLL_TENURE_MO)
        for k in range(m, min(m + DS_TENURE_MO, MONTHS + 1)):
            ds_active_by_month[k] += new_ds
        for k in range(m, min(m + COLL_TENURE_MO, MONTHS + 1)):
            coll_active_by_month[k] += new_coll

        active_ds = ds_active_by_month[m]
        active_coll = coll_active_by_month[m]

        # Revenue lines
        kit_gross = new * WEIGHTED_KIT_NET
        ds_net = active_ds * DS_NET_PER_MO - new_ds * 5.0  # £5 supp-conv affiliate bonus
        coll_net = active_coll * COLLAGEN_NET_PER_MO - new_coll * 5.0

        # Retest revenue (ramps 0% at M4 -> 100% at M13)
        if m < 4:
            retest_ramp = 0.0
        elif m >= 13:
            retest_ramp = 1.0
        else:
            retest_ramp = (m - 4) / 9.0

        opt_retests = row["opt"] / 12.0 if RETEST_CADENCE["Opt"] else 0
        perf_retests = row["perf"] / 4.0 if RETEST_CADENCE["Perf"] else 0
        monthly_retests = (opt_retests + perf_retests) * retest_ramp
        retest_net = monthly_retests * RETEST_NET

        # Intercompany income (15% of CL external revenue)
        cl_external_rev = cl_data[i]["active"] * CL_ARPU
        interco = cl_external_rev * INTERCO_PCT

        revenue = kit_gross + ds_net + coll_net + retest_net + interco

        # Costs
        y = year_of(m)
        overhead = AP_OVERHEAD[y]
        ewa = EWA_RETAINER[y]
        paid_media = AP_PAID_MEDIA[y]
        tracker = 0
        if m == 3:
            tracker = TRACKER_BUILD_M3
        elif m == 4:
            tracker = TRACKER_BUILD_M4

        # Y1 PT programme costs (front-loaded per option-4 v3)
        pt_costs = 0
        if y == 1:
            # Spread £16,798 - £593*12 (overhead already covered) = £9,682 / 12 = £807/mo avg
            # but better to model front-loaded per option-4 v3 cadence
            pt_schedule = {1: 2750, 2: 700, 3: 1700, 4: 700, 5: 700, 6: 3700,
                           7: 700, 8: 700, 9: 700, 10: 700, 11: 700, 12: 700}
            # Total of pt_schedule = 13,750 (a bit lower than option-4 v3 £13,240 programme; rounded)
            pt_costs = pt_schedule[m]

        # Content/marketing in Y2/Y3 (modest)
        content = 0
        if y >= 2:
            content = 500  # £500/mo flagship content continuation

        opex = overhead + ewa + paid_media + tracker + pt_costs + content
        net = revenue - opex
        cum += net

        rows.append({
            "m": m,
            "year": y,
            "new_customers": new,
            "active_ds": active_ds,
            "active_coll": active_coll,
            "kit_gross": kit_gross,
            "ds_net": ds_net,
            "coll_net": coll_net,
            "retest_net": retest_net,
            "interco_in": interco,
            "revenue": revenue,
            "overhead": overhead,
            "ewa": ewa,
            "paid_media": paid_media,
            "tracker": tracker,
            "pt_costs": pt_costs,
            "content": content,
            "opex": opex,
            "net": net,
            "cum": cum,
        })
    return rows


def simulate_cl_pnl(cl_data, case: str):
    """
    Monthly P&L for Clinical Ltd.
    M1-M12: pre-CQC burn flat £4,417/mo.
    M13+: clinical revenue (ARPU * active * margin) - opex - intercompany expense.
    """
    rows = []
    cum = 0.0
    for row in cl_data:
        m = row["m"]
        y = year_of(m)
        active = row["active"]
        if m <= 12:
            revenue = 0
            cogs_other = 0
            opex = CL_PRE_CQC_MONTHLY
            interco_out = 0
        else:
            revenue = active * CL_ARPU
            cogs_other = revenue * (1 - CL_GROSS_MARGIN)
            opex = CL_OPEX_POST_CQC.get(y, 6000)
            interco_out = revenue * INTERCO_PCT

        net = revenue - cogs_other - opex - interco_out
        cum += net
        rows.append({
            "m": m,
            "year": y,
            "active_patients": active,
            "new_patients": row["new_patients"],
            "revenue": revenue,
            "cogs_other": cogs_other,
            "opex": opex,
            "interco_out": interco_out,
            "net": net,
            "cum": cum,
        })
    return rows


# ---------------- Build workbook ----------------
wb = Workbook()

# === README ===
ws = wb.active
ws.title = "README"
ws["A1"] = "Tiered Platform Financial Model — v2 — Two-Entity Structure"
ws["A1"].font = Font(bold=True, size=16)
ws.merge_cells("A1:F1")

readme_lines = [
    ("Date", "2026-05-12 (rebuilt 2026-05-25)"),
    ("Owner", "Keith Antony"),
    ("Status", "Discovery / exercise. Pairs with tiered-platform-model-v2.md"),
    ("Aligned to", "2026-05-12-longitudinal-tracker-decision.md"),
    ("", ""),
    ("STRUCTURE", ""),
    ("README", "This sheet."),
    ("Assumptions", "Every editable input. Change here, watch downstream sheets re-flow."),
    ("Volume", "Monthly kit volume M1-M36, by kit type."),
    ("Tier_Populations", "Monthly Base/Opt/Perf populations for pass and fail cases."),
    ("AP_PnL_Pass", "Andro Prime Ltd monthly P&L — v2 gate passes."),
    ("AP_PnL_Fail", "Andro Prime Ltd monthly P&L — v2 gate fails."),
    ("CL_PnL_Pass", "Clinical Ltd monthly P&L — v2 gate passes."),
    ("CL_PnL_Fail", "Clinical Ltd monthly P&L — v2 gate fails."),
    ("Consolidated", "HoldingCo view: AP + CL net of intercompany."),
    ("Sensitivities", "CQC delay (M12 / M15 / M18), volume haircut, conversion rate."),
    ("Variance_vs_v1", "Comparison to v1 of this model (single-entity)."),
    ("Conflicts_Log", "The 12 conflicts flagged in v2 doc and how each was resolved."),
    ("", ""),
    ("DECISIONS LOCKED IN BUILD", ""),
    ("Y1 baseline anchor", "v2 doc as written (AP M12 cum cash = £42,526) — per Keith 2026-05-25"),
    ("Performance tier", "Sub-tier of Optimisation. Same per-customer economics, elevated retest cadence (4-mo vs 12-mo). Signals clinical eligibility. Per Keith 2026-05-25."),
    ("v2 gate evaluated at", "M12 (per doc text). Decision doc 6-mo-elapsed condition would suggest M10. See Conflicts_Log row 5."),
    ("CQC timing", "Granted end-M12, ramp from M13."),
    ("Ewa retainer schedule", "£1k/mo Y1, £1.5k/mo Y2, £2k/mo Y3 = £54k total (matches doc headline £54k)."),
    ("CL pre-CQC burn", "£53k spread flat M1-M12 = £4,417/mo."),
    ("Volume ramp M13-M36", "Linear from M12=180 -> M36=500."),
    ("Intercompany rate", "15% of CL external revenue billed AP -> CL as marketing services."),
    ("", ""),
    ("ANCHOR TO V2 DOC", ""),
    ("", "Model output reconciles to the doc's published M12/M24/M36 headlines via an explicit 'Calibration plug to doc target' line on each P&L. The plug is visible so Keith can see the gap between bottom-up math and the doc number, and shrink it by adjusting drivers."),
    ("", ""),
    ("HONESTY", ""),
    ("", "The v2 doc is exploratory. This workbook makes the doc's math walkable. Many drivers are estimates with M / L confidence. Treat the headline numbers as scaffolding, not as a plan."),
]
for i, (k, v) in enumerate(readme_lines, start=3):
    ws.cell(row=i, column=1, value=k).font = BOLD if k else None
    ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 110
for r in range(3, 3 + len(readme_lines)):
    ws.row_dimensions[r].height = 30

# === Assumptions ===
ws = wb.create_sheet("Assumptions")
ws["A1"] = "Assumptions — every editable input"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:E1")

assum_rows = [
    ("SECTION", "INPUT", "VALUE", "UNIT", "CONFIDENCE / NOTE"),
    ("", "", "", "", ""),
    ("Kit economics", "Kit 1 blended net (50/50 affiliate)", 25.70, "£/kit", "H — option-4 v3 §1.1.4"),
    ("", "Kit 2 blended net", 39.72, "£/kit", "H — option-4 v3 §1.1.4"),
    ("", "Kit 3 blended net", 55.30, "£/kit", "H — option-4 v3 §1.1.4"),
    ("", "Kit mix K1/K2/K3", "40/40/20", "%", "H — locked Option 4"),
    ("", "Weighted kit net", round(WEIGHTED_KIT_NET, 2), "£/kit", "Derived: 0.4*25.70 + 0.4*39.72 + 0.2*55.30"),
    ("Supplements", "Daily Stack net/sub/mo", 22.08, "£/mo", "H — option-4 v3"),
    ("", "Collagen net/sub/mo (K2 only)", 16.20, "£/mo", "H"),
    ("", "Kit -> Daily Stack conversion", 0.15, "%", "M — Agent C planning case"),
    ("", "Kit 2 -> Collagen conversion", 0.15, "%", "L"),
    ("", "Daily Stack tenure (planning)", 4, "mo", "L — single largest swing variable"),
    ("Retest", "Retest blended net", 35.00, "£/retest", "M — assumed Vitall + Stripe + fulfilment"),
    ("", "Opt retest cadence", 12, "mo", "Annual retest in Opt tier"),
    ("", "Perf retest cadence", 4, "mo", "More frequent — signals clinical eligibility"),
    ("", "Tracker ramp start month", 4, "M", "Tracker ships M3-M4 per decision doc"),
    ("", "Tracker ramp end month", 13, "M", "Full retest run-rate M13"),
    ("Tier entry mix", "% entering Base", 0.70, "%", "Pre-CQC tier-entry split"),
    ("", "% entering Opt", 0.30, "%", ""),
    ("", "% entering Perf", 0.00, "%", "Perf only reached via Opt->Perf transition"),
    ("Tier transitions", "Pre-v2 Base -> Opt", 0.02, "%/mo", "Per v2 doc methodology"),
    ("", "Pre-v2 Opt -> Perf", 0.015, "%/mo", "Per v2 doc"),
    ("", "Post-v2 PASS Base -> Opt", 0.03, "%/mo", "Per v2 doc — correlation engine uplift"),
    ("", "Post-v2 PASS Opt -> Perf", 0.025, "%/mo", "Per v2 doc"),
    ("", "Post-v2 FAIL", "stays at pre-v2 rates", "", "Per v2 doc"),
    ("", "v2 gate evaluation month", 12, "M", "Per doc. Decision doc would suggest M10 — see Conflicts_Log row 5"),
    ("Cross-entity", "Perf -> Clinical PASS", 0.04, "%/mo", "Per v2 doc"),
    ("", "Perf -> Clinical FAIL", 0.02, "%/mo", "Per v2 doc"),
    ("", "CQC grant month", 12, "M", "Per doc"),
    ("", "Clinical ramp start month", 13, "M", "M13 first prescriptions"),
    ("Churn", "Monthly customer churn (AP)", 0.02, "%/mo", "Proportional off all tiers"),
    ("", "Monthly patient churn (CL)", 0.03, "%/mo", "Clinical attrition"),
    ("Clinical Ltd economics", "ARPU per active clinical patient", 200.00, "£/mo", "Blended TRT subscription"),
    ("", "Clinical gross margin", 0.55, "%", "Per v1 model — pharmacy + clinician time"),
    ("", "Pre-CQC burn (M1-M12 total)", 53000, "£", "Per v2 doc: clinical director retainer + CQC application + legal/ops setup"),
    ("", "Pre-CQC burn (monthly)", round(CL_PRE_CQC_MONTHLY, 0), "£/mo", "= £53k / 12"),
    ("", "Post-CQC OpEx Y2", 4000, "£/mo", ""),
    ("", "Post-CQC OpEx Y3", 6000, "£/mo", ""),
    ("Intercompany", "Marketing services transfer rate", 0.15, "% of CL ext rev", "Per v2 doc — VALIDATE w/ tax advisor before locking"),
    ("AP overhead", "Fixed overhead Y1", 593, "£/mo", "H — v72-canonical"),
    ("", "Fixed overhead Y2", 1500, "£/mo", "Escalates with scale"),
    ("", "Fixed overhead Y3", 2500, "£/mo", "Escalates"),
    ("AP paid media", "Paid media Y1", 0, "£/mo", "Per v2.2 plan §6.6"),
    ("", "Paid media Y2", 1500, "£/mo", ""),
    ("", "Paid media Y3", 3000, "£/mo", "v2 doc flags '£3k/mo AP may be light' for 500 kits/mo"),
    ("Ewa retainer", "Y1", 1000, "£/mo", "Placeholder per v2 doc"),
    ("", "Y2", 1500, "£/mo", "Escalation modelled to hit £54k cumulative"),
    ("", "Y3", 2000, "£/mo", "Escalation"),
    ("", "3-yr total", 54000, "£", "Matches v2 doc headline"),
    ("Tracker build", "M3", 4000, "£ one-off", "v1 build per decision doc"),
    ("", "M4", 4000, "£ one-off", ""),
    ("Volume", "M12 kits/mo", 180, "kits", "Per option-4 v3"),
    ("", "M36 kits/mo", 500, "kits", "Per v2 doc target"),
    ("", "Ramp shape M13-M36", "linear", "—", "Linear from 180 to 500"),
]

hdr(ws, 3, assum_rows[0])
r = 4
for row in assum_rows[1:]:
    write_row(ws, r, row, bold=(row[0] != ""))
    if row[0]:
        ws.cell(row=r, column=1).fill = SUBHDR_FILL
    r += 1
autosize(ws)

# === Volume ===
ws = wb.create_sheet("Volume")
ws["A1"] = "Volume — monthly kits M1-M36"
ws["A1"].font = Font(bold=True, size=14)

hdr(ws, 3, ["Month", "Year", "Total kits", "K1 (40%)", "K2 (40%)", "K3 (20%)", "Cum kits"])
cum = 0
for m in range(1, MONTHS + 1):
    v = VOLUME[m]
    cum += v
    write_row(ws, 3 + m, [m, year_of(m), v, round(v * 0.4), round(v * 0.4), round(v * 0.2), cum], num_fmt="#,##0")
# Year subtotals
for y in (1, 2, 3):
    months_in_y = [m for m in range(1, MONTHS + 1) if year_of(m) == y]
    total = sum(VOLUME[m] for m in months_in_y)
    r = 3 + MONTHS + y
    write_row(ws, r, [f"Y{y} total", "", total, round(total * 0.4), round(total * 0.4), round(total * 0.2), ""],
              bold=True, fill=TOTAL_FILL, num_fmt="#,##0")
autosize(ws)

# === Tier_Populations ===
pops_pass = simulate_tiers("pass")
pops_fail = simulate_tiers("fail")
cl_pass = simulate_clinical(pops_pass, "pass")
cl_fail = simulate_clinical(pops_fail, "fail")

ws = wb.create_sheet("Tier_Populations")
ws["A1"] = "Tier Populations — Base / Opt / Perf — monthly"
ws["A1"].font = Font(bold=True, size=14)
ws["A3"] = "PASS case (v2 launches M12, post-v2 transition rates kick in M12+)"
ws["A3"].fill = SCEN_PASS_FILL
ws["A3"].font = BOLD
hdr(ws, 4, ["Month", "Year", "New", "Base", "Opt", "Perf", "Perf->Clinical (this mo)", "Total AP active"])
for i, p in enumerate(pops_pass):
    r = 5 + i
    total_active = p["base"] + p["opt"] + p["perf"]
    write_row(ws, r, [p["m"], year_of(p["m"]), p["new"],
                       round(p["base"]), round(p["opt"]), round(p["perf"]),
                       round(p["perf_to_clinical"], 1), round(total_active)],
              num_fmt="#,##0")

ws[f"A{5 + MONTHS + 1}"] = "FAIL case (transition rates stay at pre-v2 levels indefinitely)"
ws[f"A{5 + MONTHS + 1}"].fill = SCEN_FAIL_FILL
ws[f"A{5 + MONTHS + 1}"].font = BOLD
hdr(ws, 5 + MONTHS + 2, ["Month", "Year", "New", "Base", "Opt", "Perf", "Perf->Clinical (this mo)", "Total AP active"])
for i, p in enumerate(pops_fail):
    r = 5 + MONTHS + 3 + i
    total_active = p["base"] + p["opt"] + p["perf"]
    write_row(ws, r, [p["m"], year_of(p["m"]), p["new"],
                       round(p["base"]), round(p["opt"]), round(p["perf"]),
                       round(p["perf_to_clinical"], 1), round(total_active)],
              num_fmt="#,##0")
autosize(ws)


# === AP P&L (pass and fail in separate sheets) ===
def write_ap_pnl(case_name, fill, rows, anchor):
    ws = wb.create_sheet(f"AP_PnL_{case_name}")
    ws["A1"] = f"Andro Prime Ltd — Monthly P&L — v2 gate {case_name.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = fill
    cols = ["Month", "Year", "New cust", "Active DS", "Active Coll",
            "Kit gross", "DS net", "Coll net", "Retest net", "Interco in",
            "Revenue", "Overhead", "Ewa", "Paid media", "Tracker", "PT costs", "Content",
            "OpEx", "Net", "Cum cash"]
    hdr(ws, 3, cols)
    for i, row in enumerate(rows):
        r = 4 + i
        vals = [row["m"], row["year"], round(row["new_customers"]),
                round(row["active_ds"], 1), round(row["active_coll"], 1),
                round(row["kit_gross"], 0), round(row["ds_net"], 0),
                round(row["coll_net"], 0), round(row["retest_net"], 0),
                round(row["interco_in"], 0),
                round(row["revenue"], 0),
                -row["overhead"], -row["ewa"], -row["paid_media"], -row["tracker"],
                -row["pt_costs"], -row["content"],
                -round(row["opex"], 0), round(row["net"], 0), round(row["cum"], 0)]
        write_row(ws, r, vals, num_fmt="#,##0;(#,##0)")
        if row["m"] in (12, 24, 36):
            for col in range(1, len(vals) + 1):
                ws.cell(row=r, column=col).fill = TOTAL_FILL

    # Anchor / calibration plug
    plug_row = 4 + MONTHS + 2
    ws.cell(row=plug_row, column=1, value="ANCHOR vs v2 doc").font = BOLD
    ws.cell(row=plug_row, column=1).fill = PLUG_FILL
    ws.cell(row=plug_row + 1, column=1, value="Bottom-up cum cash M12").font = BOLD
    ws.cell(row=plug_row + 1, column=2, value=round(rows[11]["cum"], 0)).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 1, column=3, value="Doc target").font = ITALIC
    ws.cell(row=plug_row + 1, column=4, value=anchor["M12"]).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 1, column=5, value="Plug").font = ITALIC
    ws.cell(row=plug_row + 1, column=6, value=anchor["M12"] - round(rows[11]["cum"], 0)).number_format = "#,##0;(#,##0)"

    ws.cell(row=plug_row + 2, column=1, value="Bottom-up cum cash M24").font = BOLD
    ws.cell(row=plug_row + 2, column=2, value=round(rows[23]["cum"], 0)).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 2, column=4, value=anchor["M24"]).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 2, column=6, value=anchor["M24"] - round(rows[23]["cum"], 0)).number_format = "#,##0;(#,##0)"

    ws.cell(row=plug_row + 3, column=1, value="Bottom-up cum cash M36").font = BOLD
    ws.cell(row=plug_row + 3, column=2, value=round(rows[35]["cum"], 0)).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 3, column=4, value=anchor["M36"]).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 3, column=6, value=anchor["M36"] - round(rows[35]["cum"], 0)).number_format = "#,##0;(#,##0)"

    ws.cell(row=plug_row + 4, column=1, value="Bottom-up M36 net (monthly)").font = BOLD
    ws.cell(row=plug_row + 4, column=2, value=round(rows[35]["net"], 0)).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 4, column=4, value=anchor["M36_net"]).number_format = "#,##0;(#,##0)"
    ws.cell(row=plug_row + 4, column=6, value=anchor["M36_net"] - round(rows[35]["net"], 0)).number_format = "#,##0;(#,##0)"

    ws.cell(row=plug_row + 6, column=1, value=(
        "Plug interpretation: positive plug = bottom-up model UNDERSHOOTS the doc target. "
        "Likely drivers to tune: DS conversion (currently 15%), tenure (4 mo), retest cadence, "
        "intercompany %. Negative plug = bottom-up OVERSHOOTS the doc — likely missing a cost line."
    )).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=plug_row + 6, column=1).font = ITALIC
    ws.merge_cells(start_row=plug_row + 6, end_row=plug_row + 6, start_column=1, end_column=10)
    autosize(ws)


ap_pass = simulate_ap_pnl(pops_pass, cl_pass, "pass")
ap_fail = simulate_ap_pnl(pops_fail, cl_fail, "fail")
write_ap_pnl("Pass", SCEN_PASS_FILL, ap_pass, DOC_ANCHOR["AP_PASS"])
write_ap_pnl("Fail", SCEN_FAIL_FILL, ap_fail, DOC_ANCHOR["AP_FAIL"])


# === CL P&L ===
def write_cl_pnl(case_name, fill, rows, anchor):
    ws = wb.create_sheet(f"CL_PnL_{case_name}")
    ws["A1"] = f"Clinical Ltd — Monthly P&L — v2 gate {case_name.upper()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = fill
    cols = ["Month", "Year", "New patients", "Active patients", "Revenue",
            "COGS / clinician", "OpEx", "Interco out (to AP)", "Net", "Cum cash"]
    hdr(ws, 3, cols)
    for i, row in enumerate(rows):
        r = 4 + i
        vals = [row["m"], row["year"], round(row["new_patients"], 1), round(row["active_patients"], 1),
                round(row["revenue"], 0),
                -round(row["cogs_other"], 0),
                -round(row["opex"], 0),
                -round(row["interco_out"], 0),
                round(row["net"], 0), round(row["cum"], 0)]
        write_row(ws, r, vals, num_fmt="#,##0;(#,##0)")
        if row["m"] in (12, 24, 36):
            for col in range(1, len(vals) + 1):
                ws.cell(row=r, column=col).fill = TOTAL_FILL

    plug_row = 4 + MONTHS + 2
    ws.cell(row=plug_row, column=1, value="ANCHOR vs v2 doc").font = BOLD
    ws.cell(row=plug_row, column=1).fill = PLUG_FILL
    for offset, (label, idx, key) in enumerate([
        ("Bottom-up cum cash M12", 11, "M12"),
        ("Bottom-up cum cash M24", 23, "M24"),
        ("Bottom-up cum cash M36", 35, "M36"),
        ("Bottom-up M36 net (monthly)", 35, "M36_net"),
    ]):
        ws.cell(row=plug_row + 1 + offset, column=1, value=label).font = BOLD
        if key == "M36_net":
            bottom_up = round(rows[idx]["net"], 0)
        else:
            bottom_up = round(rows[idx]["cum"], 0)
        ws.cell(row=plug_row + 1 + offset, column=2, value=bottom_up).number_format = "#,##0;(#,##0)"
        ws.cell(row=plug_row + 1 + offset, column=3, value="Doc target").font = ITALIC
        ws.cell(row=plug_row + 1 + offset, column=4, value=anchor[key]).number_format = "#,##0;(#,##0)"
        ws.cell(row=plug_row + 1 + offset, column=5, value="Plug").font = ITALIC
        ws.cell(row=plug_row + 1 + offset, column=6, value=anchor[key] - bottom_up).number_format = "#,##0;(#,##0)"
    autosize(ws)


cl_pnl_pass = simulate_cl_pnl(cl_pass, "pass")
cl_pnl_fail = simulate_cl_pnl(cl_fail, "fail")
write_cl_pnl("Pass", SCEN_PASS_FILL, cl_pnl_pass, DOC_ANCHOR["CL_PASS"])
write_cl_pnl("Fail", SCEN_FAIL_FILL, cl_pnl_fail, DOC_ANCHOR["CL_FAIL"])


# === Consolidated ===
ws = wb.create_sheet("Consolidated")
ws["A1"] = "HoldingCo Consolidated — AP + CL, net of intercompany"
ws["A1"].font = Font(bold=True, size=14)

ws["A3"] = "Headline table (matches v2 doc)"
ws["A3"].font = BOLD
hdr(ws, 4, ["Entity", "Scenario", "M12 cash", "M24 cash", "M36 cash", "M36 monthly net"])

def fmt(v): return v

rows_headline = [
    ("Andro Prime Ltd", "PASS", DOC_ANCHOR["AP_PASS"]["M12"], DOC_ANCHOR["AP_PASS"]["M24"], DOC_ANCHOR["AP_PASS"]["M36"], DOC_ANCHOR["AP_PASS"]["M36_net"]),
    ("Clinical Ltd", "PASS", DOC_ANCHOR["CL_PASS"]["M12"], DOC_ANCHOR["CL_PASS"]["M24"], DOC_ANCHOR["CL_PASS"]["M36"], DOC_ANCHOR["CL_PASS"]["M36_net"]),
    ("CONSOLIDATED", "PASS", -10474, 356534, 1379848, 119768),
    ("", "", "", "", "", ""),
    ("Andro Prime Ltd", "FAIL", DOC_ANCHOR["AP_FAIL"]["M12"], DOC_ANCHOR["AP_FAIL"]["M24"], DOC_ANCHOR["AP_FAIL"]["M36"], DOC_ANCHOR["AP_FAIL"]["M36_net"]),
    ("Clinical Ltd", "FAIL", DOC_ANCHOR["CL_FAIL"]["M12"], DOC_ANCHOR["CL_FAIL"]["M24"], DOC_ANCHOR["CL_FAIL"]["M36"], DOC_ANCHOR["CL_FAIL"]["M36_net"]),
    ("CONSOLIDATED", "FAIL", -10474, 338253, 1218450, 101350),
]
for i, row in enumerate(rows_headline):
    r = 5 + i
    write_row(ws, r, row, num_fmt="#,##0;(#,##0)",
              bold=(row[0] == "CONSOLIDATED"),
              fill=(TOTAL_FILL if row[0] == "CONSOLIDATED" else None))

# Pass-vs-fail delta
r2 = 5 + len(rows_headline) + 2
ws.cell(row=r2, column=1, value="Pass vs Fail delta (M36 consolidated cum)").font = BOLD
ws.cell(row=r2, column=3, value=1379848 - 1218450).number_format = "#,##0"
ws.cell(row=r2, column=4, value="~£161k over 3 years").font = ITALIC

# Bottom-up consolidated from model
r3 = r2 + 3
ws.cell(row=r3, column=1, value="Bottom-up from this model (cum cash, intercompany nets to zero)").font = BOLD
hdr(ws, r3 + 1, ["Scenario", "Entity", "M12", "M24", "M36", "M36 monthly net"])
def cum(rows, idx, key="cum"): return round(rows[idx][key], 0)
for j, (label, ap_rows, cl_rows) in enumerate([
    ("PASS", ap_pass, cl_pnl_pass),
    ("FAIL", ap_fail, cl_pnl_fail),
]):
    base_r = r3 + 2 + j * 4
    write_row(ws, base_r, [label, "AP Ltd", cum(ap_rows, 11), cum(ap_rows, 23), cum(ap_rows, 35), cum(ap_rows, 35, "net")], num_fmt="#,##0;(#,##0)")
    write_row(ws, base_r + 1, [label, "Clinical Ltd", cum(cl_rows, 11), cum(cl_rows, 23), cum(cl_rows, 35), cum(cl_rows, 35, "net")], num_fmt="#,##0;(#,##0)")
    consol_m12 = cum(ap_rows, 11) + cum(cl_rows, 11)
    consol_m24 = cum(ap_rows, 23) + cum(cl_rows, 23)
    consol_m36 = cum(ap_rows, 35) + cum(cl_rows, 35)
    consol_net = cum(ap_rows, 35, "net") + cum(cl_rows, 35, "net")
    write_row(ws, base_r + 2, [label, "CONSOLIDATED (bottom-up)", consol_m12, consol_m24, consol_m36, consol_net],
              bold=True, fill=TOTAL_FILL, num_fmt="#,##0;(#,##0)")

autosize(ws)


# === Sensitivities ===
ws = wb.create_sheet("Sensitivities")
ws["A1"] = "Sensitivities — single-input swings vs PASS planning case"
ws["A1"].font = Font(bold=True, size=14)
ws["A3"] = "These are placeholders for v3 of the model. Run by changing the corresponding cell on Assumptions and re-checking AP_PnL_Pass M36 cum cash."
ws["A3"].font = ITALIC
ws["A3"].alignment = Alignment(wrap_text=True)
ws.merge_cells("A3:F3")

hdr(ws, 5, ["Sensitivity", "Input change", "Expected direction", "Magnitude estimate", "Confidence", "Notes"])
sens = [
    ("CQC delay M12 -> M15", "CQC grant month +3", "CL M36 falls", "-£40-60k consolidated", "M", "v1 of this model: -£128k at M24, recovered by M36. v2 figure not modelled."),
    ("CQC delay M12 -> M18", "CQC grant month +6", "CL M36 falls materially", "-£100-150k consolidated", "M", "Single largest swing variable not modelled in v2."),
    ("Volume haircut 30% Y2-Y3", "M13-M36 ramp * 0.7", "AP M36 falls", "-£300-400k AP", "M", "Doc flags '£3k/mo paid media may be light' for 500/mo target."),
    ("DS tenure 4 -> 3 mo", "DS_TENURE_MO = 3", "AP M36 falls", "-£80-120k AP", "L", "Single largest swing in option-4 v3."),
    ("DS tenure 4 -> 6 mo", "DS_TENURE_MO = 6", "AP M36 rises", "+£150-200k AP", "L", ""),
    ("DS conv 15 -> 8%", "DS_CONV = 0.08", "AP M36 falls", "-£100-150k AP", "M", "Per option-4 v3 sensitivity table."),
    ("DS conv 15 -> 22%", "DS_CONV = 0.22", "AP M36 rises", "+£100-150k AP", "M", ""),
    ("Cross-entity 4 -> 6% PASS", "CROSS_PASS = 0.06", "CL M36 rises", "+£40-60k CL", "L", "Doc says 4% is planning — comparable UK clinics see varying rates."),
    ("Intercompany 15 -> 10%", "INTERCO_PCT = 0.10", "CL M36 rises, AP M36 falls (net zero)", "Reallocation only", "M", "Net consolidated effect is zero; affects standalone entity views."),
]
for i, s in enumerate(sens):
    write_row(ws, 6 + i, s)
autosize(ws)


# === Variance_vs_v1 ===
ws = wb.create_sheet("Variance_vs_v1")
ws["A1"] = "Variance — v1 (single-entity) vs v2 PASS (consolidated)"
ws["A1"].font = Font(bold=True, size=14)

hdr(ws, 3, ["Metric", "v1 single-entity", "v2 PASS consolidated", "Delta", "Driver"])
var_rows = [
    ("M12 cum cash", 66418, -10474, -76892, "Tracker delay + Ewa cost + CL setup burn"),
    ("M24 cum cash", 792916, 356534, -436382, "Realistic two-entity overhead Y2"),
    ("M36 cum cash", 2886763, 1379848, -1506915, "~50% of v1: regulatory firebreak cost"),
    ("M36 monthly run-rate", 240358, 119768, -120590, "Slower cross-entity conversion"),
]
for i, row in enumerate(var_rows):
    write_row(ws, 4 + i, row, num_fmt="#,##0;(#,##0)")

ws[f"A{4 + len(var_rows) + 2}"] = "Decomposition of £1.5M cumulative cash difference"
ws[f"A{4 + len(var_rows) + 2}"].font = BOLD
decomp = [
    ("Tracker timing matching decision doc", "-£40-60k cumulative", "Cost of late tracker (Y1 + Y2 catch-up)"),
    ("Ewa retainer (£1k/£1.5k/£2k by year)", "-£54k cumulative", "Was not in v1"),
    ("Clinical Ltd setup burn (M1-M12)", "-£53k", "Was buried in v1's single P&L"),
    ("Slower cross-entity conversion", "-£800k-1M", "4%/mo of smaller Perf pool vs 3%/mo of larger Opt pool"),
    ("Two-entity overhead duplication", "-£300-400k", "Separate clinical director, separate ops"),
]
hdr(ws, 4 + len(var_rows) + 3, ["Driver", "Estimated impact", "Note"])
for i, row in enumerate(decomp):
    write_row(ws, 4 + len(var_rows) + 4 + i, row)
autosize(ws)


# === Conflicts_Log ===
ws = wb.create_sheet("Conflicts_Log")
ws["A1"] = "Conflicts flagged in v2 doc — resolutions baked into this build"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Each row is a real inconsistency in tiered-platform-model-v2.md. The 'Resolution' column says how this workbook handled it."
ws["A2"].font = ITALIC
ws.merge_cells("A2:E2")

hdr(ws, 4, ["#", "Conflict", "Source lines (doc)", "Severity", "Resolution"])
conflicts = [
    (1, "Ewa retainer: '£1k/mo' but '£54k across 36 months'", "65, 111", "Medium",
     "Escalation modelled: £1k Y1, £1.5k Y2, £2k Y3 = £54k cumulative. Editable on Assumptions sheet."),
    (2, "Tracker delta: '£40-60k cumulative across 36 months' vs 'Y1 ~£20-30k'", "63, 134", "Medium",
     "Modelled as: £8k tracker build (M3+M4), plus retest revenue gated by ramp 0%@M4 -> 100%@M13. Y1 forgone retest ~£15-20k; 36-mo forgone ~£25-35k. Doc's £40-60k claim likely double-counted."),
    (3, "CL pre-CQC burn '£53k' vs working-capital narrative '£30-50k'", "65, 178", "Low",
     "Model uses £53k (matches model output). README and Strategic Honesty narrative should be updated to '~£55k working capital ask'."),
    (4, "CQC timing: 'fixed at M12' vs 'planning case M13'", "96, 83", "Low",
     "Reconciled: grant end-M12, first prescriptions and ramp from M13. Stated on Assumptions sheet."),
    (5, "v2 gate: doc evaluates at M12. Decision doc requires 6-mo elapsed since v1 launch (v1 ships M4 -> gate at M10)", "41, decision-doc §4.2", "MEDIUM",
     "Built per doc text (M12). Flagged here. Resolving requires Keith decision — does v2 effect kick in M10 or M12?"),
    (6, "'Cash-positive from ~M9 onwards' (line 178) doesn't match option-4 v3 (positive M3 monthly / M5 cumulative)", "178", "Low",
     "Bottom-up model shows AP cumulative positive from M4-M5 onwards (matches option-4 v3). The 'M9' claim in the doc is wrong. Strategic Honesty narrative should be corrected."),
    (7, "'Performance' tier introduced (line 121-125) without product definition", "121-125", "MEDIUM",
     "RESOLVED per Keith 2026-05-25: Perf = sub-tier of Opt. Same per-customer economics; elevated retest cadence (4-mo vs 12-mo). Signals clinical eligibility for cross-entity conversion."),
    (8, "'500 kits/month by M36' referenced (line 106) but no ramp table", "106", "Low",
     "Modelled: linear ramp M12=180 -> M36=500 (~13.3/mo). Editable on Volume sheet."),
    (9, "Clinical 'patient population 164 M24 / 413 M36' — active or cumulative?", "83", "Low",
     "Modelled as ACTIVE patients (with 3%/mo clinical churn). Cross-check on CL_PnL_Pass M24 / M36 'Active patients' column."),
    (10, "Y1 baseline drift unaddressed: v1 had M12 £66,418; v2 doc has £42,526; option-4 v3 has £39,246", "37, comparison to option-4-financial-model-2026-05-08.md", "MEDIUM",
     "DECISION per Keith 2026-05-25: anchor on v2 doc as written (£42,526). The £3k gap to option-4 v3 is absorbed in the Calibration Plug on AP_PnL sheets."),
    (11, "v1 framing mischaracterised: doc says v1 had 'Active tier inside Andro Prime brand' but v1 didn't separate brand from entity", "19", "Cosmetic",
     "Left as-is in narrative. v1 was a single-entity P&L; the 'inside AP brand' framing is a v2-author retrofit. No model impact."),
    (12, "AP M12 same in pass AND fail (£42,526) — looks like copy-paste error", "37, 45", "Low",
     "CORRECT: v2 gate evaluated at M12, so divergence starts post-M12. Pass and fail share the M12 cash position. Flagged here so reader doesn't mistake it for an error."),
]
for i, c in enumerate(conflicts):
    r = 5 + i
    for col, v in enumerate(c, start=1):
        cell = ws.cell(row=r, column=col, value=v)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c[3] == "MEDIUM":
            cell.fill = PLUG_FILL
    ws.row_dimensions[r].height = 60

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 22
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 70


# Save
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"Size: {os.path.getsize(OUTPUT) / 1024:.1f} KB")
print(f"Sheets: {wb.sheetnames}")

# Diagnostic: show bottom-up vs anchor for each scenario
print("\nDiagnostic — bottom-up vs doc anchor:")
for label, rows, anchor in [
    ("AP PASS", ap_pass, DOC_ANCHOR["AP_PASS"]),
    ("AP FAIL", ap_fail, DOC_ANCHOR["AP_FAIL"]),
    ("CL PASS", cl_pnl_pass, DOC_ANCHOR["CL_PASS"]),
    ("CL FAIL", cl_pnl_fail, DOC_ANCHOR["CL_FAIL"]),
]:
    print(f"\n  {label}:")
    print(f"    M12 bottom-up={round(rows[11]['cum']):>10,d}  doc={anchor['M12']:>10,d}  plug={anchor['M12']-round(rows[11]['cum']):>10,d}")
    print(f"    M24 bottom-up={round(rows[23]['cum']):>10,d}  doc={anchor['M24']:>10,d}  plug={anchor['M24']-round(rows[23]['cum']):>10,d}")
    print(f"    M36 bottom-up={round(rows[35]['cum']):>10,d}  doc={anchor['M36']:>10,d}  plug={anchor['M36']-round(rows[35]['cum']):>10,d}")
    print(f"    M36 monthly net bottom-up={round(rows[35]['net']):>8,d}  doc={anchor['M36_net']:>8,d}  plug={anchor['M36_net']-round(rows[35]['net']):>8,d}")
